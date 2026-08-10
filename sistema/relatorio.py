### FUNÇÕES PARA GERAÇÃO DE RELATÓRIOS ###

# Importações
import pandas as pd
from datetime import datetime
from copy import copy
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.shapes import GraphicalProperties
import sistema.tratamento_dados as f
import streamlit as st

# RELATÓRIO SEM AGRUPAMENTO

# Criação do Dataframe
def criar_relatorio_conciliação_simples(
    df_conciliado,
    saldo_inicial,  
    mov_extrato, 
    mov_controle,
    total_extrato,
    total_controle,
    nome_usuario
):
    """
    Cria um relatório completo da conciliação em um DataFrame estruturado
    que será exportado para excel
    """
    
    relatorio_dados = []
    
    # CABEÇALHO DO RELATÓRIO
    
    relatorio_dados.append(["RELATÓRIO DE CONCILIAÇÃO BANCÁRIA"])
    relatorio_dados.append(["Transações Conciliadas"])
    relatorio_dados.append([]) # Linha em branco
    relatorio_dados.append(["Usuário Responsável:", nome_usuario])
    relatorio_dados.append(["Data da Realização:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    relatorio_dados.append([])  # Linha em branco
    
    # DADOS GERAIS DA CONCILIAÇÃO (CONCILIADAS)
    relatorio_dados.append(["RESUMO GERAL DA CONCILIAÇÃO"])
    relatorio_dados.append(["Indicador", "Extrato", "Controle Financeiro", "Diferença"])
    relatorio_dados.append(["Total de Movimentações", mov_extrato, mov_controle, "=B9-C9"])
    relatorio_dados.append(["Saldo Inicial (R$)", saldo_inicial, saldo_inicial, "=B10-C10"])
    relatorio_dados.append(["Valor Total Movimentado (R$)", total_extrato, total_controle, "=B11-C11"])
    relatorio_dados.append(["Saldo Final (R$):", st.session_state.saldo_final_ex, st.session_state.saldo_final_cf, "=B12-C12"])
    relatorio_dados.append([])  # Linha em branco

    operacoes_divergentes = df_conciliado[df_conciliado['status_conciliacao'] == 'NÃO CONCILIADO']
    operacoes_convergentes = df_conciliado[df_conciliado['status_conciliacao'] == 'CONCILIADA']
    
    relatorio_div = relatorio_dados.copy()
    relatorio_conv = relatorio_dados.copy()
    
    # Relatório das Operações Não Conciliadas
    if len(operacoes_divergentes) > 0:
        relatorio_div.append(["OPERAÇÕES DIVERGENTES NO EXTRATO (NÃO CONCILIADAS)"])
        
        # Cabeçalho das operações divergentes
        cabecalho_extrato = [
            "Data", "Documento", "Descrição", "Valor (R$)", "Crítica"
        ]
        
        cabecalho_controle = [
            "Data", "Descrição", "Contraparte", "Plano de Contas", "Valor (R$)", "Crítica"
        ]
        
        # Adiciona dados que tem no extrato e não estão no controle financeiro
        extrato_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_extrato,
                'documento': row.get('documento_extrato', ''),
                'descricao': row.get('descricao_extrato', ''),
                'valor': float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "vazio"
            }
            if linha['valor'] != "vazio":
                extrato_divergente.append(linha)
        
        extrato_divergente = f.ordenar_por_data_br(extrato_divergente)
        sum_divergente_extrato = len(extrato_divergente)
        total_divergente_extrato = 0
        for i in extrato_divergente:
            total_divergente_extrato += i['valor']
        
        relatorio_div.append(["Total de Transações Não Conciliadas:", "", "", sum_divergente_extrato])
        relatorio_div.append(["Valor Total (R$):", "", "", total_divergente_extrato])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho_extrato)
        
        for i in extrato_divergente:
            relatorio_div.append([i['data'], i['documento'], i['descricao'], i['valor'], ''])
            
        ultima_linha_extrato = 19 +len(extrato_divergente)-1
        
        relatorio_div.append(["", "", "TOTAL", f"=SUM(D19:D{ultima_linha_extrato})"])
        
        relatorio_div.append([])  # Linha em branco
            
        # Adiciona dados que tem no controle financeiro e não estão no extrato
        linha_controle = ultima_linha_extrato + 3
        
        relatorio_div.append(["OPERAÇÕES DIVERGENTES NO CONTROLE FINANCEIRO (NÃO CONCILIADAS)"])
        
        linha_controle += 1
        
        controle_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_controle,
                'recurso': row.get('recurso_controle', ''),
                'contraparte': row.get('contraparte_controle', ''),
                'plano de contas': row.get('plano de contas_controle'),
                'valor': float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "vazio"
            }
            if linha['valor'] != "vazio":
                controle_divergente.append(linha)
            
        controle_divergente = f.ordenar_por_data_br(controle_divergente)
        sum_divergente_controle = len(controle_divergente)
        total_divergente_controle = 0
        for i in controle_divergente:
            total_divergente_controle += i['valor']
    
        relatorio_div.append(["Total de Transações Não Conciliadas:", "", "", sum_divergente_controle])
        relatorio_div.append(["Valor Total (R$):", "", "", total_divergente_controle])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho_controle)
        linha_controle += 4
        
        primeira_linha_controle = linha_controle
        for i in controle_divergente:
                relatorio_div.append([i['data'], i['recurso'], i['contraparte'], i['plano de contas'], i['valor'], ''])
        
        ultima_linha_controle = primeira_linha_controle +len(controle_divergente)-1
        relatorio_div.append(["", "", "TOTAL", "", f"=SUM(E{primeira_linha_controle}:E{ultima_linha_controle})"])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(["Documento gerado automaticamente pelo sistema de conciliação bancária."])
            
    else:
        relatorio_div.append(["NENHUMA OPERAÇÃO DIVERGENTE ENCONTRADA"])
        
    # Relatório das Operações Conciliadas
    if len(operacoes_convergentes) > 0:
        sum_convergentes = len(operacoes_convergentes)
        total_convergentes = sum(operacoes_convergentes['valor_extrato'])
        
        relatorio_conv.append(["OPERAÇÕES CONVERGENTES (CONCILIADAS)"])
        relatorio_conv.append(["Total de Transações Conciliadas:", "", "", "", "", sum_convergentes])
        relatorio_conv.append(["Valor Total Conciliado (R$):", "", "", "", "", total_convergentes])
        relatorio_conv.append([])  # Linha em branco
        
        # Cabeçalho das operações convergentes
        cabecalho = [
            "Data Extrato", "Documento Extrato", "Descrição Extrato", "Valor Extrato (R$)", 
            "Data Controle", "Descrição Controle", "Contraparte Controle", "Plano de Contas Controle", "Valor Controle (R$)"
        ]
        relatorio_conv.append(cabecalho)
        
        # Adiciona cada operação convergente
        for _, row in operacoes_convergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = [
                data_extrato,
                row.get('documento_extrato', ''),
                row.get('descricao_extrato', ''),
                float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "",
                data_controle,
                row.get('recurso_controle', ''),
                row.get('contraparte_controle', ''),
                row.get('plano de contas_controle'),
                float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "",
            ]
            relatorio_conv.append(linha)
        
        
        ultima_linha = 19 +len(operacoes_convergentes)-1
        
        relatorio_conv.append(["", "", "TOTAL", f"=SUM(D19:D{ultima_linha})", "", "", "", "", f"=SUM(I19:I{ultima_linha})"])
        relatorio_conv.append([])  # Linha em branco
        relatorio_conv.append(["Documento gerado automaticamente pelo sistema de conciliação bancária."])
    else:
        relatorio_conv.append(["NENHUMA OPERAÇÃO CONCILIADA ENCONTRADA"])
    
    # Resumo Gráfico da Conciliação
    relatorio_grafico = []
    relatorio_grafico.append(["RESUMO EXECUTIVO\nAUDITORIA DE CONCILIAÇÃO BANCÁRIA"])
    relatorio_grafico.append([]) # Linha em branco
    relatorio_grafico.append([]) # Linha em branco
    relatorio_grafico.append(["Responsável:", nome_usuario])
    relatorio_grafico.append(["Data da Realização:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    relatorio_grafico.append([])  # Linha em branco
    relatorio_grafico.append(["INDICADORES GERAIS"])
    relatorio_grafico.append(["Saldo Inicial (Extrato):", "", "='Transações Conciliadas'!B10"])
    relatorio_grafico.append(["Saldo Final (Extrato):", "", "='Transações Conciliadas'!B12"])
    relatorio_grafico.append(["Transferência entre Contas (Extrato):","", 0])
    relatorio_grafico.append(["Fluxo de caixa (Controle):","", "='Transações Conciliadas'!C11"])
    relatorio_grafico.append(["Diferença de Saldo (Extrato − Controle):", "", "=C9-C10-C11"])
    relatorio_grafico.append(["Acuracidade", "", '=IFERROR(C11/(C9-C10),"")'])
    for i in range(17):
        relatorio_grafico.append([])  # Linhas em branco para espaçamento
    relatorio_grafico.append(["Dados dos Gráficos (escrever aqui)"])
    relatorio_grafico.append(["Origem", "Conciliadas", "Não Conciliadas"])
    relatorio_grafico.append(["Extrato", 90, 80])
    relatorio_grafico.append(["Controle Financeiro", 60, 50])
    
    # Converte para DataFrame
    df_relatorio_div = pd.DataFrame(relatorio_div)
    df_relatorio_conv = pd.DataFrame(relatorio_conv)
    df_relatorio_grafico = pd.DataFrame(relatorio_grafico)
    
    return df_relatorio_conv, df_relatorio_div, df_relatorio_grafico

# Formatação do Relatório Excel
def exportar_relatorio_excel_simples(df_relatorio_conv, df_relatorio_div, df_relatorio_grafico):
    """
    Exporta o relatório para Excel.
    """

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Escreve os DataFrames
        
        df_relatorio_grafico.to_excel(
            writer,
            sheet_name="Resumo Executivo",
            index=False,
            header=False,
        )
        
        df_relatorio_conv.to_excel(
            writer,
            sheet_name="Transações Conciliadas",
            index=False,
            header=False,
        )

        df_relatorio_div.to_excel(
            writer,
            sheet_name="Transações Não Conciliadas",
            index=False,
            header=False,
        )

        # Obtém as abas
        ws_relatorio_grafico = writer.book["Resumo Executivo"]
        ws_conciliadas = writer.book["Transações Conciliadas"]
        ws_nao_conciliadas = writer.book["Transações Não Conciliadas"]

        # Aplica a formatação
        formatar_aba_grafico(ws_relatorio_grafico)
        formatar_aba_conciliadas(ws_conciliadas)
        formatar_aba_nao_conciliadas(ws_nao_conciliadas)
        
        # Ajuste personalizado das larguras
        largura_conciliadas = {"A": 45.86, "B": 21.29, "C": 43.43, "D": 19.57, "E": 16.29, "F": 25.57, "G": 27.14, "H": 57.00, "I": 21.00}
        largura_nao_conciliadas = {"A": 45.86, "B": 15.00, "C": 45.29, "D": 31.57, "E": 13.29, "F": 21.29}
        largura_relatorio_grafico = {"A": 31.86, "B": 16.29, "C": 18.57, "D": 18.57}
        altura_linhas = {1: 31.5, 2: 19.5, 7: 21.75, 8: 27.75}
        altura_grafico = {1: 60, 2: 21.75, 3: 7.5, 4: 15, 5: 15, 6:7.5, 7: 21.75, 8: 24, 9: 24, 10:24, 11: 24, 12: 24, 13: 24, 15: 20, 16: 20,
                          17: 14.25, 18:14.25, 19:14.25, 20:14.25, 21:14.25, 22:14.25, 23:14.25, 24:14.25, 25:14.25, 26:14.25,
                          27:14.25, 28:14.25, 29:14.25, 30:14.25, 31:15, 32:15, 33:14.25, 34:14.25}
        
        ajustar_largura_colunas(ws_conciliadas, largura_conciliadas)
        ajustar_largura_colunas(ws_nao_conciliadas, largura_nao_conciliadas)
        ajustar_altura_linhas(ws_conciliadas, altura_linhas)
        ajustar_altura_linhas(ws_nao_conciliadas, altura_linhas)
        ajustar_largura_colunas(ws_relatorio_grafico, largura_relatorio_grafico)
        ajustar_altura_linhas(ws_relatorio_grafico,altura_grafico)
        
    output.seek(0)
    return output.getvalue()

# Helpers formatação do relatório

def formatar_aba_conciliadas(ws):
    # Fontes
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Calibri', size=12, bold=True, color='1f3864')
    user_data_font = Font(name='Calibri', size=10, bold=True, color='000000')
    resumo_geral_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    topo_tabela_font = Font(name='Calibri', size=10.5, bold=True, color='FFFFFF')
    tabela_dados_font = Font(name='Calibri', size=10, color='000000')
    tabela_titulo_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    total_valor_font = Font(name='Calibri', size=11, bold=True, color='2e7d32')
    total_valor_dados_font = Font(name='Calibri', size=13, bold=True, color='2e7d32')
    linha_final_font = Font(name='Calibri', size=8, italic=True, color='808080')
    
    # Preenchimentos
    azul_escuro = PatternFill(start_color='1f3864', end_color='1f3864', fill_type='solid')
    azul_claro = PatternFill(start_color='d9e2f3', end_color='d9e2f3', fill_type='solid')
    azul_medio = PatternFill(start_color='2e75b6', end_color='2e75b6', fill_type='solid')
    verde_escuro = PatternFill(start_color='2e7d32', end_color='2e7d32', fill_type='solid')
    verde_claro = PatternFill(start_color='e2f0d9', end_color='e2f0d9', fill_type='solid')
    
    # Bordas
    borda_fina = Side(border_style='thin', color='D0D0D0')
    borda_branca = Side(border_style='thin', color='FFFFFF')
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    
    # Alinhamentos
    centralizado = Alignment(horizontal='center', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Formatos numéricos
    formato_moeda = '[Black]"R$" #,##0.00;[Red]-"R$" #,##0.00'
    
    #Remover Linhas de Grade
    ws.sheet_view.showGridLines = False

    # ==================================================
    # LINHA 1 - TÍTULO
    # ==================================================

    ws.merge_cells("A1:I1")

    cell = ws["A1"]

    cell.font = titulo_font
    cell.fill = azul_escuro
    cell.alignment = centralizado

    # ==================================================
    # LINHA 2 - SUBTÍTULO
    # ==================================================

    ws.merge_cells("A2:I2")

    cell = ws["A2"]

    cell.font = subtitulo_font
    cell.fill = azul_claro
    cell.alignment = centralizado

    # ==================================================
    # DADOS DO RELATÓRIO
    # ==================================================

    ws["A4"].font = user_data_font
    ws["B4"].font = tabela_dados_font

    ws["A5"].font = user_data_font
    ws["B5"].font = tabela_dados_font
    
    # ==================================================
    # RESUMO GERAL
    # ==================================================
    
    ws.merge_cells("A7:I7")
    
    cell = ws["A7"]
    
    cell.font = resumo_geral_font
    cell.fill = azul_medio
    cell.alignment = centralizado
    
    # CABEÇALHO DA TABELA
    for cell in ws["A8:D8"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    # PRIMEIRA LINHA DA TABELA
    for cell in ws["A9:D9"][0]:
        cell.font = tabela_dados_font
        cell.border = borda
        
    ws["B9"].alignment = alinhamento_direita
    ws["C9"].alignment = alinhamento_direita
    ws["D9"].alignment = alinhamento_direita
    
    # DEMAIS LINHAS DA TABELA
    for linha in range (10, 13):
        for cell in ws[f"A{linha}:D{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda

        for cell in ws[f"B{linha}:D{linha}"][0]:
            cell.alignment = alinhamento_direita
            cell.number_format = formato_moeda
    
    # ==================================================
    # CONCILIADAS
    # ==================================================
    
    # Título da Tabela
    ws.merge_cells("A14:I14")
    ws["A14"].font = tabela_titulo_font
    ws["A14"].fill = verde_escuro
    ws["A14"].alignment = centralizado
    
    # Informações da Conciliação
    # Total de Transações
    ws.merge_cells("A15:E15")
    ws["A15"].font = total_valor_font
    ws["A15"].fill = verde_claro
    ws["A15"].border = borda
    
    ws.merge_cells("F15:I15")
    ws["F15"].font = total_valor_dados_font
    ws["F15"].alignment = alinhamento_direita
    ws["F15"].fill = verde_claro
    ws["F15"].border = borda
    
    # Valor Total Conciliado
    ws.merge_cells("A16:E16")
    ws["A16"].font = total_valor_font
    ws["A16"].fill = verde_claro
    ws["A16"].border = borda
    
    ws.merge_cells("F16:I16")
    ws["F16"].font = total_valor_dados_font
    ws["F16"].alignment = alinhamento_direita
    ws["F16"].number_format = formato_moeda
    ws["F16"].fill = verde_claro
    ws["F16"].border = borda
    
    # ==================================================
    # TABELA CONCILIADAS
    # ==================================================
    
    for cell in ws["A18:I18"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    ultima_linha = ws.max_row - 3
    
    for linha in range(19, ultima_linha + 1):
        for cell in ws[f"A{linha}:I{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda
    
        ws[f"B{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].number_format = formato_moeda
        ws[f"I{linha}"].alignment = alinhamento_direita
        ws[f"I{linha}"].number_format = formato_moeda
    
    linha_total = ws.max_row -2
    
    for cell in ws[f"A{linha_total}:I{linha_total}"][0]:
        cell.fill = verde_claro
        cell.font = user_data_font
        cell.border = borda
        
    ws[f"C{linha_total}"].alignment = alinhamento_direita
    ws[f"C{linha_total}"].font = user_data_font
    ws[f"D{linha_total}"].number_format = formato_moeda
    ws[f"D{linha_total}"].alignment = alinhamento_direita
    ws[f"D{linha_total}"].font = user_data_font
    ws[f"I{linha_total}"].number_format = formato_moeda
    ws[f"I{linha_total}"].alignment = alinhamento_direita
    ws[f"I{linha_total}"].font = user_data_font
    
    linha_final = ws.max_row
    ws[f"A{linha_final}"].font = linha_final_font
    
def formatar_aba_nao_conciliadas(ws):
    # Fontes
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Calibri', size=12, bold=True, color='c0392b')
    user_data_font = Font(name='Calibri', size=10, bold=True, color='000000')
    resumo_geral_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    topo_tabela_font = Font(name='Calibri', size=10.5, bold=True, color='FFFFFF')
    tabela_dados_font = Font(name='Calibri', size=10, color='000000')
    tabela_titulo_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    total_valor_font = Font(name='Calibri', size=11, bold=True, color='c0392b')
    total_valor_dados_font = Font(name='Calibri', size=13, bold=True, color='c0392b')
    linha_final_font = Font(name='Calibri', size=8, italic=True, color='808080')
    
    # Preenchimentos
    azul_escuro = PatternFill(start_color='1f3864', end_color='1f3864', fill_type='solid')
    azul_medio = PatternFill(start_color='2e75b6', end_color='2e75b6', fill_type='solid')
    vermelho_escuro = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
    vermelho_claro = PatternFill(start_color='fbe1dd', end_color='fbe1dd', fill_type='solid')
    
    # Bordas
    fina = Side(border_style='thin', color='D0D0D0')
    branca= Side(border_style='thin', color='FFFFFF')
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    borda_branca = Border(right=branca)
    
    # Alinhamentos
    centralizado = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Formatos numéricos
    formato_moeda = '[Black]"R$" #,##0.00;[Red]-"R$" #,##0.00'
    
    #Remover Linhas de Grade
    ws.sheet_view.showGridLines = False

    # Encontrando onde começam as divergentes do controle
    linha_atual = None
    for row in ws.iter_rows():
        valor = row[0].value

        if valor == "OPERAÇÕES DIVERGENTES NO CONTROLE FINANCEIRO (NÃO CONCILIADAS)":
            linha_atual = row[0].row
            break


    # ==================================================
    # LINHA 1 - TÍTULO
    # ==================================================
    
    ws.merge_cells("A1:F1")
    
    cell = ws["A1"]
    
    cell.font = titulo_font
    cell.fill = azul_escuro
    cell.alignment = centralizado
    
    # ==================================================
    # LINHA 2 - SUBTÍTULO
    # ==================================================
    
    ws.merge_cells("A2:F2")
    
    cell = ws["A2"]
    
    cell.font = subtitulo_font
    cell.fill = vermelho_claro
    cell.alignment = centralizado
    
    # ==================================================
    # DADOS DO RELATÓRIO
    # ==================================================
    
    ws["A4"].font = user_data_font
    ws["B4"].font = tabela_dados_font
    
    ws["A5"].font = user_data_font
    ws["B5"].font = tabela_dados_font
    
    # ==================================================
    # RESUMO GERAL
    # ==================================================
    
    ws.merge_cells("A7:F7")
    
    cell = ws["A7"]
    
    cell.font = resumo_geral_font
    cell.fill = azul_medio
    cell.alignment = centralizado
    
    # CABEÇALHO DA TABELA
    ws.merge_cells("D8:F8")
    
    for cell in ws["A8:D8"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    # PRIMEIRA LINHA DA TABELA
    ws.merge_cells("D9:F9")
    
    for cell in ws["A9:D9"][0]:
        cell.font = tabela_dados_font
        cell.border = borda
    
    ws["B9"].alignment = alinhamento_direita
    ws["C9"].alignment = alinhamento_direita
    ws["D9"].alignment = alinhamento_direita
    
    # DEMAIS LINHAS DA TABELA
    for linha in range (10, 13):
        ws.merge_cells(f"D{linha}:F{linha}")
        
        for cell in ws[f"A{linha}:D{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda
    
        for cell in ws[f"B{linha}:D{linha}"][0]:
            cell.alignment = alinhamento_direita
            cell.number_format = formato_moeda
    
    # ==================================================
    # NÃO CONCILIADAS EXTRATO
    # ==================================================
    
    # Título da Tabela
    ws.merge_cells("A14:F14")
    ws["A14"].font = tabela_titulo_font
    ws["A14"].fill = vermelho_escuro
    ws["A14"].alignment = centralizado
    
    # Informações da Conciliação
    # Total de Transações Não Conciliadas Extrato
    ws.merge_cells("A15:C15")
    ws["A15"].font = total_valor_font
    ws["A15"].fill = vermelho_claro
    ws["A15"].border = borda
    
    ws.merge_cells("D15:F15")
    ws["D15"].font = total_valor_dados_font
    ws["D15"].alignment = alinhamento_direita
    ws["D15"].fill = vermelho_claro
    ws["D15"].border = borda
    
    # Valor Total Não Conciliado Extrato
    ws.merge_cells("A16:C16")
    ws["A16"].font = total_valor_font
    ws["A16"].fill = vermelho_claro
    ws["A16"].border = borda
    
    ws.merge_cells("D16:F16")
    ws["D16"].font = total_valor_dados_font
    ws["D16"].alignment = alinhamento_direita
    ws["D16"].number_format = formato_moeda
    ws["D16"].fill = vermelho_claro
    ws["D16"].border = borda
    
    # ==================================================
    # TABELA NÃO CONCILIADAS EXTRATO
    # ==================================================
    
    ws.merge_cells("E18:F18")
    
    for cell in ws["A18:E18"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    ultima_linha = linha_atual -2
    
    for linha in range(19, ultima_linha + 1):
        ws.merge_cells(f"E{linha}:F{linha}")
        for cell in ws[f"A{linha}:E{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda
    
        ws[f"B{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].number_format = formato_moeda
    
    linha_total_extrato = linha_atual - 2
    
    for cell in ws[f"A{linha_total_extrato}:E{linha_total_extrato}"][0]:
        cell.fill = vermelho_claro
        cell.font = user_data_font
        cell.border = borda
    
    ws[f"C{linha_total_extrato}"].alignment = alinhamento_direita
    ws[f"C{linha_total_extrato}"].font = user_data_font
    ws[f"D{linha_total_extrato}"].number_format = formato_moeda
    ws[f"D{linha_total_extrato}"].alignment = alinhamento_direita
    ws[f"D{linha_total_extrato}"].font = user_data_font
    
    # ==================================================
    # NÃO CONCILIADAS CONTROLE
    # ==================================================
    
    if linha_atual is None:
        raise ValueError("Não foi encontrada a seção de operações divergentes do controle.")

    # Título da Tabela
    ws.merge_cells(f"A{linha_atual}:F{linha_atual}")
    ws[f"A{linha_atual}"].font = tabela_titulo_font
    ws[f"A{linha_atual}"].fill = vermelho_escuro
    ws[f"A{linha_atual}"].alignment = centralizado
    
    # Informações da Conciliação
    # Total de Transações Não Conciliadas Extrato
    ws.merge_cells(f"A{linha_atual + 1}:C{linha_atual + 1}")
    ws[f"A{linha_atual + 1}"].font = total_valor_font
    ws[f"A{linha_atual + 1}"].fill = vermelho_claro
    ws[f"A{linha_atual + 1}"].border = borda
    
    ws.merge_cells(f"D{linha_atual + 1}:F{linha_atual + 1}")
    ws[f"D{linha_atual + 1}"].font = total_valor_dados_font
    ws[f"D{linha_atual + 1}"].alignment = alinhamento_direita
    ws[f"D{linha_atual + 1}"].fill = vermelho_claro
    ws[f"D{linha_atual + 1}"].border = borda
    
    # Valor Total Não Conciliado Extrato
    ws.merge_cells(f"A{linha_atual + 2}:C{linha_atual + 2}")
    ws[f"A{linha_atual + 2}"].font = total_valor_font
    ws[f"A{linha_atual + 2}"].fill = vermelho_claro
    ws[f"A{linha_atual + 2}"].border = borda
    
    ws.merge_cells(f"D{linha_atual + 2}:F{linha_atual + 2}")
    ws[f"D{linha_atual + 2}"].font = total_valor_dados_font
    ws[f"D{linha_atual + 2}"].alignment = alinhamento_direita
    ws[f"D{linha_atual + 2}"].number_format = formato_moeda
    ws[f"D{linha_atual + 2}"].fill = vermelho_claro
    ws[f"D{linha_atual + 2}"].border = borda
    
    # ==================================================
    # TABELA NÃO CONCILIADAS CONTROLE
    # ==================================================
    
    
    for cell in ws[f"A{linha_atual + 4}:F{linha_atual + 4}"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    ultima_linha = ws.max_row - 3
    
    for linha in range(linha_atual + 5, ultima_linha + 1):
        for cell in ws[f"A{linha}:E{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda
    
        ws[f"B{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].number_format = formato_moeda
    
    linha_total_controle = ws.max_row - 2
    
    for cell in ws[f"A{linha_total_controle}:F{linha_total_controle}"][0]:
        cell.fill = vermelho_claro
        cell.font = user_data_font
        cell.border = borda
    
    ws[f"C{linha_total_controle}"].font = user_data_font
    ws[f"C{linha_total_controle}"].alignment = alinhamento_direita
    ws[f"E{linha_total_controle}"].font = user_data_font
    ws[f"E{linha_total_controle}"].number_format = formato_moeda
    ws[f"E{linha_total_controle}"].alignment = alinhamento_direita
    
    linha_final = ws.max_row
    ws[f"A{linha_final}"].font = linha_final_font

def formatar_aba_grafico(ws):
    # Fontes
    titulo_font = Font(name='Calibri', size=14, bold=True, color='000000')
    user_data_font = Font(name='Calibri', size=10, bold=True, color='000000')
    usuario_data_font = Font(name='Calibri', size=10, color='000000')
    resumo_geral_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    tabela_font = Font(name='Calibri', size=11, bold=True, color='000000')
    tabela_dados_font = Font(name='Calibri', size=13, bold=True, color='000000')
    tabelinha_font = Font(name='Calibri', size=11, bold=True, color='808080')
    tabelinha_topo_font = Font(name='Calibri', size=9, bold=True, color='000000')
    tabelinha_dados = Font(name='Calibri', size=11, color='000000')
    diferenca_font = Font(name='Calibri', size=11, bold=True, color='b8860b')
    
    # Preenchimentos
    verde_agua = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
    
    # Bordas
    fina = Side(border_style='thin', color='D0D0D0')
    vermelha= Side(border_style='thick', color='c0392b')
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    borda_vermelha = Border(top=vermelha)
    
    # Alinhamentos
    centralizado = Alignment(horizontal='center', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Formatos numéricos
    formato_moeda = '[Black]"R$" #,##0.00;[Red]-"R$" #,##0.00'
    formato_porcentagem = "0.00%"
    
    #Remover Linhas de Grade
    ws.sheet_view.showGridLines = False
    
    # Criação do Gráfico
    categorias = Reference(
        ws,
        min_col = 1,
        min_row= 33,
        max_row = 34
    )
    dados = Reference(
        ws,
        min_col= 2,
        max_col=3,
        min_row=32,
        max_row=34
    )
    grafico = BarChart()
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.x_axis.delete = False
    grafico.title = "Transações Conciliadas x Não Conciliadas"
    grafico.series[0].graphicalProperties.solidFill = "4A7729"  # Verde
    grafico.series[1].graphicalProperties.solidFill = "B01116"  # Vermelho
    
    # ==================================================
    # LOGOTIPO E TÍTULO
    # ==================================================
    
    logo = Image("sistema/logo_semfundo.png")
    logo.width = 120
    logo.height = 70
    ws.add_image(logo, "A1")
    ws.merge_cells("A1:D1")
    ws["A1"].font = titulo_font
    ws["A1"].alignment = centralizado
    ws.merge_cells("A2:D2")
    ws.border = borda_vermelha
    
    # ==================================================
    # DADOS GERAIS
    # ==================================================
    ws["A4"].font = user_data_font
    ws["A5"].font = user_data_font
    ws["B4"].font = usuario_data_font
    ws["B5"].font = usuario_data_font
    
    # ==================================================
    # INDICADORES
    # ==================================================
    ws.merge_cells("A7:D7")
    ws["A7"].font = resumo_geral_font
    ws["A7"].fill = verde_agua
    ws["A7"].alignment = centralizado
    
    for linha in range(8, 12):
        ws.merge_cells(f"A{linha}:B{linha}")
        ws[f"A{linha}"].font = tabela_font
        ws[f"A{linha}"].border = borda
    
    ws.merge_cells("A12:B12")
    ws["A12"].font = diferenca_font
    ws["A12"].border = borda
    for linha in range(13, 15):
        ws.merge_cells(f"A{linha}:B{linha}")
        ws[f"A{linha}"].font = tabela_font
        ws[f"A{linha}"].border = borda

    for linha in range(8, 13):
        ws.merge_cells(f"C{linha}:D{linha}")
        ws[f"C{linha}"].font = tabela_dados_font
        ws[f"C{linha}"].border = borda
        ws[f"C{linha}"].number_format = formato_moeda
        ws[f"C{linha}"].alignment = centralizado
    
    ws.merge_cells("C13:D13")
    ws["C13"].font = tabela_dados_font
    ws["C13"].border = borda
    ws["C13"].alignment = centralizado
    ws["C13"].number_format = formato_porcentagem
    
    ws.add_chart(grafico, "A15")
    
    ws["A31"].font = tabelinha_font
    for cell in ws["A32:C32"][0]:
        cell.font = tabelinha_topo_font
    
    for linha in range(33, 35):
        for cell in ws[f"A{linha}:C{linha}"][0]:
            cell.font = tabelinha_dados
    
    for linha in range(33, 35):
            for cell in ws[f"B{linha}:C{linha}"][0]:
                cell.alignment = alinhamento_direita        

def ajustar_largura_colunas(ws, larguras):
    """
    Ajusta a largura das colunas conforme o dicionário informado.

    Exemplo:
        {"A": 15, "B": 20, "C": 35}
    """

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

def ajustar_altura_linhas(ws, alturas):
    """
    Ajusta a altura das linhas.

    Exemplo:
        {1: 30, 2: 22, 7: 25}
    """
    for linha, altura in alturas.items():
        ws.row_dimensions[linha].height = altura

# RELATÓRIO COM AGRUPAMENTO

# Criação do Dataframe
def criar_relatorio_conciliação_agrupamento(
    df_conciliado,
    saldo_inicial,  
    mov_extrato, 
    mov_controle,
    total_extrato,
    total_controle,
    nome_usuario
):
    """
    Cria um relatório completo da conciliação em um DataFrame estruturado
    que será exportado para excel
    """
    
    relatorio_dados = []
    
    # CABEÇALHO DO RELATÓRIO
    
    relatorio_dados.append(["RELATÓRIO DE CONCILIAÇÃO BANCÁRIA"])
    relatorio_dados.append(["Transações Conciliadas"])
    relatorio_dados.append([]) # Linha em branco
    relatorio_dados.append(["Usuário Responsável:", nome_usuario])
    relatorio_dados.append(["Data da Realização:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    relatorio_dados.append([])  # Linha em branco
    
    # DADOS GERAIS DA CONCILIAÇÃO (CONCILIADAS)
    relatorio_dados.append(["RESUMO GERAL DA CONCILIAÇÃO"])
    relatorio_dados.append(["Indicador", "Extrato", "Controle Financeiro", "Diferença"])
    relatorio_dados.append(["Total de Movimentações", mov_extrato, mov_controle, "=B9-C9"])
    relatorio_dados.append(["Saldo Inicial (R$)", saldo_inicial, saldo_inicial, "=B10-C10"])
    relatorio_dados.append(["Valor Total Movimentado (R$)", total_extrato, total_controle, "=B11-C11"])
    relatorio_dados.append(["Saldo Final (R$):", st.session_state.saldo_final_ex, st.session_state.saldo_final_cf, "=B12-C12"])
    relatorio_dados.append([])  # Linha em branco

    operacoes_divergentes = df_conciliado[df_conciliado['status_conciliacao'] == 'NÃO CONCILIADO']
    operacoes_convergentes_simples = df_conciliado[(df_conciliado['status_conciliacao'] == 'CONCILIADA') &
                                           (df_conciliado["tipo_conciliacao"] != "AGRUPAMENTO")]
    operacoes_convergentes_agrupadas = df_conciliado[(df_conciliado['status_conciliacao'] == 'CONCILIADA') &
                                               (df_conciliado["tipo_conciliacao"] == "AGRUPAMENTO")]
    
    relatorio_div = relatorio_dados.copy()
    relatorio_conv = relatorio_dados.copy()
    
    # Relatório das Operações Não Conciliadas
    if len(operacoes_divergentes) > 0:
        relatorio_div.append(["OPERAÇÕES DIVERGENTES NO EXTRATO (NÃO CONCILIADAS)"])
        
        # Cabeçalho das operações divergentes
        cabecalho_extrato = [
            "Data", "Documento", "Descrição", "Valor (R$)", "Crítica"
        ]
        
        cabecalho_controle = [
            "Data", "Descrição", "Contraparte", "Plano de Contas", "Valor (R$)", "Crítica"
        ]
        
        # Adiciona dados que tem no extrato e não estão no controle financeiro
        extrato_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_extrato,
                'documento': row.get('documento_extrato', ''),
                'descricao': row.get('descricao_extrato', ''),
                'valor': float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "vazio"
            }
            if linha['valor'] != "vazio":
                extrato_divergente.append(linha)
        
        extrato_divergente = f.ordenar_por_data_br(extrato_divergente)
        sum_divergente_extrato = len(extrato_divergente)
        total_divergente_extrato = 0
        for i in extrato_divergente:
            total_divergente_extrato += i['valor']
        
        relatorio_div.append(["Total de Transações Não Conciliadas:", "", "", sum_divergente_extrato])
        relatorio_div.append(["Valor Total (R$):", "", "", total_divergente_extrato])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho_extrato)
        
        for i in extrato_divergente:
            relatorio_div.append([i['data'], i['documento'], i['descricao'], i['valor'], ''])
            
        ultima_linha_extrato = 19 +len(extrato_divergente)-1
        
        relatorio_div.append(["", "", "TOTAL", f"=SUM(D19:D{ultima_linha_extrato})"])
        
        relatorio_div.append([])  # Linha em branco
            
        # Adiciona dados que tem no controle financeiro e não estão no extrato
        linha_controle = ultima_linha_extrato + 3
        
        relatorio_div.append(["OPERAÇÕES DIVERGENTES NO CONTROLE FINANCEIRO (NÃO CONCILIADAS)"])
        
        linha_controle += 1
        
        controle_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_controle,
                'recurso': row.get('recurso_controle', ''),
                'contraparte': row.get('contraparte_controle', ''),
                'plano de contas': row.get('plano de contas_controle'),
                'valor': float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "vazio"
            }
            if linha['valor'] != "vazio":
                controle_divergente.append(linha)
            
        controle_divergente = f.ordenar_por_data_br(controle_divergente)
        sum_divergente_controle = len(controle_divergente)
        total_divergente_controle = 0
        for i in controle_divergente:
            total_divergente_controle += i['valor']
    
        relatorio_div.append(["Total de Transações Não Conciliadas:", "", "", sum_divergente_controle])
        relatorio_div.append(["Valor Total (R$):", "", "", total_divergente_controle])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho_controle)
        linha_controle += 4
        
        primeira_linha_controle = linha_controle
        for i in controle_divergente:
                relatorio_div.append([i['data'], i['recurso'], i['contraparte'], i['plano de contas'], i['valor'], ''])
        
        ultima_linha_controle = primeira_linha_controle +len(controle_divergente)-1
        relatorio_div.append(["", "", "TOTAL", "", f"=SUM(E{primeira_linha_controle}:E{ultima_linha_controle})"])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(["Documento gerado automaticamente pelo sistema de conciliação bancária."])
            
    else:
        relatorio_div.append(["NENHUMA OPERAÇÃO DIVERGENTE ENCONTRADA"])
        
    # Relatório das Operações Conciliadas
    if len(operacoes_convergentes_simples) > 0:
        sum_convergentes = len(operacoes_convergentes_simples)
        total_convergentes = sum(operacoes_convergentes_simples['valor_extrato'])
        
        relatorio_conv.append(["OPERAÇÕES CONVERGENTES (CONCILIADAS)"])
        relatorio_conv.append(["Total de Transações Conciliadas:", "", "", "", "", sum_convergentes])
        relatorio_conv.append(["Valor Total Conciliado (R$):", "", "", "", "", total_convergentes])
        relatorio_conv.append([])  # Linha em branco
        
        # Cabeçalho das operações convergentes
        cabecalho = [
            "Data Extrato", "Documento Extrato", "Descrição Extrato", "Valor Extrato (R$)", 
            "Data Controle", "Descrição Controle", "Contraparte Controle", "Plano de Contas Controle", "Valor Controle (R$)"
        ]
        
        # Cabeçalho das operações convergentes
        cabecalho_agrupamento = [
            "Agrupamento", "Data Extrato", "Documento Extrato", "Descrição Extrato", "Valor Extrato (R$)", 
            "Data Controle", "Descrição Controle", "Contraparte Controle", "Plano de Contas Controle", "Valor Controle (R$)"
        ]
        relatorio_conv.append(cabecalho)
        
        # Adiciona cada operação convergente
        for _, row in operacoes_convergentes_simples.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = [
                data_extrato,
                row.get('documento_extrato', ''),
                row.get('descricao_extrato', ''),
                float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "",
                data_controle,
                row.get('recurso_controle', ''),
                row.get('contraparte_controle', ''),
                row.get('plano de contas_controle'),
                float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "",
            ]
            relatorio_conv.append(linha)
        
        
        ultima_linha = 19 +len(operacoes_convergentes_simples)-1
        
        relatorio_conv.append(["", "", "TOTAL", f"=SUM(D19:D{ultima_linha})", "", "", "", "", f"=SUM(I19:I{ultima_linha})"])
        relatorio_conv.append([])  # Linha em branco
        
        # AGRUPAMENTOS
        sum_convergentes_agrupamentos = len(operacoes_convergentes_agrupadas)
        total_convergentes_agrupamentos = (operacoes_convergentes_agrupadas['valor_extrato'].sum())
        
        
        relatorio_conv.append(["OPERAÇÕES CONVERGENTES POR AGRUPAMENTO (CONCILIADAS)"])
        relatorio_conv.append(["Total de Transações Conciliadas:", "", "", "", "", sum_convergentes_agrupamentos])
        relatorio_conv.append(["Valor Total Conciliado (R$):", "", "", "", "", total_convergentes_agrupamentos])
        relatorio_conv.append([])  # Linha em branco
        relatorio_conv.append(cabecalho_agrupamento)
        primeira_linha_agrupamentos = ultima_linha + 8
        for id_agr, grupo in operacoes_convergentes_agrupadas.groupby("agrupamento"):

            # Separa as linhas do extrato e do controle
            extratos = (
                grupo[
                    grupo["_id_extrato"].notna()
                ]
                .sort_values("data_extrato")
                .reset_index(drop=True)
            )

            controles = (
                grupo[
                    grupo["_id_controle"].notna()
                ]
                .sort_values("data_controle")
                .reset_index(drop=True)
            )

            # Quantidade de linhas que o agrupamento ocupará
            qtd_linhas = max(len(extratos), len(controles))

            for i in range(qtd_linhas):

                ex = extratos.iloc[i] if i < len(extratos) else None
                cf = controles.iloc[i] if i < len(controles) else None

                relatorio_conv.append([

                    # ID do agrupamento aparece apenas na primeira linha
                    id_agr if i == 0 else "",

                    # -------- EXTRATO --------
                    ex.get("data_extrato", "") if ex is not None else "",
                    ex.get("documento_extrato", "") if ex is not None else "",
                    ex.get("descricao_extrato", "") if ex is not None else "",
                    float(ex["valor_extrato"]) if ex is not None else "",

                    # -------- CONTROLE --------
                    cf.get("data_controle", "") if cf is not None else "",
                    cf.get("recurso_controle", "") if cf is not None else "",
                    cf.get("contraparte_controle", "") if cf is not None else "",
                    cf.get("plano de contas_controle", "") if cf is not None else "",
                    float(cf["valor_controle"]) if cf is not None else ""

                ])
        
        ultima_linha = primeira_linha_agrupamentos + len(operacoes_convergentes_agrupadas) - 2
        relatorio_conv.append(["", "","", "TOTAL", f"=SUM(E{primeira_linha_agrupamentos}:E{ultima_linha})", "", "", "","", f"=SUM(J{primeira_linha_agrupamentos}:J{ultima_linha})"])
        relatorio_conv.append([])  # Linha em branco
        relatorio_conv.append(["Documento gerado automaticamente pelo sistema de conciliação bancária."])
    else:
        relatorio_conv.append(["NENHUMA OPERAÇÃO CONCILIADA ENCONTRADA"])
    
    # Resumo Gráfico da Conciliação
    relatorio_grafico = []
    relatorio_grafico.append(["RESUMO EXECUTIVO\nAUDITORIA DE CONCILIAÇÃO BANCÁRIA"])
    relatorio_grafico.append([]) # Linha em branco
    relatorio_grafico.append([]) # Linha em branco
    relatorio_grafico.append(["Responsável:", nome_usuario])
    relatorio_grafico.append(["Data da Realização:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    relatorio_grafico.append([])  # Linha em branco
    relatorio_grafico.append(["INDICADORES GERAIS"])
    relatorio_grafico.append(["Saldo Inicial (Extrato):", "", "='Transações Conciliadas'!B10"])
    relatorio_grafico.append(["Saldo Final (Extrato):", "", "='Transações Conciliadas'!B12"])
    relatorio_grafico.append(["Transferência entre Contas (Extrato):","", 0])
    relatorio_grafico.append(["Fluxo de caixa (Controle):","", "='Transações Conciliadas'!C11"])
    relatorio_grafico.append(["Diferença de Saldo (Extrato − Controle):", "", "=C9-C10-C11"])
    relatorio_grafico.append(["Acuracidade", "", '=IFERROR(C11/(C9-C10),"")'])
    for i in range(17):
        relatorio_grafico.append([])  # Linhas em branco para espaçamento
    relatorio_grafico.append(["Dados dos Gráficos (escrever aqui)"])
    relatorio_grafico.append(["Origem", "Conciliadas", "Não Conciliadas"])
    relatorio_grafico.append(["Extrato", 90, 80])
    relatorio_grafico.append(["Controle Financeiro", 60, 50])
    
    # Converte para DataFrame
    df_relatorio_div = pd.DataFrame(relatorio_div)
    df_relatorio_conv = pd.DataFrame(relatorio_conv)
    df_relatorio_grafico = pd.DataFrame(relatorio_grafico)
    
    return df_relatorio_conv, df_relatorio_div, df_relatorio_grafico

# Formatação do Relatório Excel
def formatar_aba_conciliadas_agrupamento(ws):
    # Fontes
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    subtitulo_font = Font(name='Calibri', size=12, bold=True, color='1f3864')
    user_data_font = Font(name='Calibri', size=10, bold=True, color='000000')
    resumo_geral_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    topo_tabela_font = Font(name='Calibri', size=10.5, bold=True, color='FFFFFF')
    tabela_dados_font = Font(name='Calibri', size=10, color='000000')
    tabela_titulo_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    total_valor_font = Font(name='Calibri', size=11, bold=True, color='2e7d32')
    total_valor_dados_font = Font(name='Calibri', size=13, bold=True, color='2e7d32')
    linha_final_font = Font(name='Calibri', size=8, italic=True, color='808080')
    
    # Preenchimentos
    azul_escuro = PatternFill(start_color='1f3864', end_color='1f3864', fill_type='solid')
    azul_claro = PatternFill(start_color='d9e2f3', end_color='d9e2f3', fill_type='solid')
    azul_medio = PatternFill(start_color='2e75b6', end_color='2e75b6', fill_type='solid')
    verde_escuro = PatternFill(start_color='2e7d32', end_color='2e7d32', fill_type='solid')
    verde_claro = PatternFill(start_color='e2f0d9', end_color='e2f0d9', fill_type='solid')
    
    # Bordas
    borda_fina = Side(border_style='thin', color='D0D0D0')
    borda_branca = Side(border_style='thin', color='FFFFFF')
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    
    # Alinhamentos
    centralizado = Alignment(horizontal='center', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Formatos numéricos
    formato_moeda = '[Black]"R$" #,##0.00;[Red]-"R$" #,##0.00'
    
    #Remover Linhas de Grade
    ws.sheet_view.showGridLines = False

    # ==================================================
    # LINHA 1 - TÍTULO
    # ==================================================

    ws.merge_cells("A1:J1")

    cell = ws["A1"]

    cell.font = titulo_font
    cell.fill = azul_escuro
    cell.alignment = centralizado

    # ==================================================
    # LINHA 2 - SUBTÍTULO
    # ==================================================

    ws.merge_cells("A2:J2")

    cell = ws["A2"]

    cell.font = subtitulo_font
    cell.fill = azul_claro
    cell.alignment = centralizado

    # ==================================================
    # DADOS DO RELATÓRIO
    # ==================================================

    ws["A4"].font = user_data_font
    ws["B4"].font = tabela_dados_font

    ws["A5"].font = user_data_font
    ws["B5"].font = tabela_dados_font
    
    # ==================================================
    # RESUMO GERAL
    # ==================================================
    
    ws.merge_cells("A7:J7")
    
    cell = ws["A7"]
    
    cell.font = resumo_geral_font
    cell.fill = azul_medio
    cell.alignment = centralizado
    
    # CABEÇALHO DA TABELA
    for cell in ws["A8:D8"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    # PRIMEIRA LINHA DA TABELA
    for cell in ws["A9:D9"][0]:
        cell.font = tabela_dados_font
        cell.border = borda
        
    ws["B9"].alignment = alinhamento_direita
    ws["C9"].alignment = alinhamento_direita
    ws["D9"].alignment = alinhamento_direita
    
    # DEMAIS LINHAS DA TABELA
    for linha in range (10, 13):
        for cell in ws[f"A{linha}:D{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda

        for cell in ws[f"B{linha}:D{linha}"][0]:
            cell.alignment = alinhamento_direita
            cell.number_format = formato_moeda
    
    # ==================================================
    # CONCILIADAS PELA CONCILIAÇÃO SIMPLES
    # ==================================================
    
    # Título da Tabela
    ws.merge_cells("A14:J14")
    ws["A14"].font = tabela_titulo_font
    ws["A14"].fill = verde_escuro
    ws["A14"].alignment = centralizado
    
    # Informações da Conciliação
    # Total de Transações
    ws.merge_cells("A15:E15")
    ws["A15"].font = total_valor_font
    ws["A15"].fill = verde_claro
    ws["A15"].border = borda
    
    ws.merge_cells("F15:J15")
    ws["F15"].font = total_valor_dados_font
    ws["F15"].alignment = alinhamento_direita
    ws["F15"].fill = verde_claro
    ws["F15"].border = borda
    
    # Valor Total Conciliado
    ws.merge_cells("A16:E16")
    ws["A16"].font = total_valor_font
    ws["A16"].fill = verde_claro
    ws["A16"].border = borda
    
    ws.merge_cells("F16:J16")
    ws["F16"].font = total_valor_dados_font
    ws["F16"].alignment = alinhamento_direita
    ws["F16"].number_format = formato_moeda
    ws["F16"].fill = verde_claro
    ws["F16"].border = borda
    
    # ==================================================
    # TABELA CONCILIADAS PELA CONCILIAÇÃO SIMPLES
    # ==================================================
    
    linha_atual = None
    for row in ws.iter_rows():
        valor = row[0].value

        if valor == "OPERAÇÕES CONVERGENTES POR AGRUPAMENTO (CONCILIADAS)":
            linha_atual = row[0].row
            break
        
    ws.merge_cells("I18:J18")    
    for cell in ws["A18:I18"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    ultima_linha = linha_atual - 3
    
    for linha in range(19, ultima_linha + 1):
        ws.merge_cells(f"I{linha}:J{linha}")
        for cell in ws[f"A{linha}:I{linha}"][0]:
                cell.font = tabela_dados_font
                cell.border = borda
    
        ws[f"B{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].alignment = alinhamento_direita
        ws[f"D{linha}"].number_format = formato_moeda
        ws[f"I{linha}"].alignment = alinhamento_direita
        ws[f"I{linha}"].number_format = formato_moeda
    
    linha_total_simples = linha_atual - 2
    
    ws.merge_cells(f"I{linha_total_simples}:J{linha_total_simples}")
    for cell in ws[f"A{linha_total_simples}:I{linha_total_simples}"][0]:
        cell.fill = verde_claro
        cell.font = user_data_font
        cell.border = borda
    
    ws[f"C{linha_total_simples}"].alignment = alinhamento_direita
    ws[f"C{linha_total_simples}"].font = user_data_font
    ws[f"D{linha_total_simples}"].number_format = formato_moeda
    ws[f"D{linha_total_simples}"].alignment = alinhamento_direita
    ws[f"D{linha_total_simples}"].font = user_data_font
    ws[f"I{linha_total_simples}"].number_format = formato_moeda
    ws[f"I{linha_total_simples}"].alignment = alinhamento_direita
    ws[f"I{linha_total_simples}"].font = user_data_font
    
    # ==================================================
    # CONCILIADAS POR AGRUPAMENTO
    # ==================================================
    
    if linha_atual is None:
        raise ValueError("Não foi encontrada a seção de conciliação por agrupamento.")
    
    # Título da Tabela
    ws.merge_cells(f"A{linha_atual}:J{linha_atual}")
    ws[f"A{linha_atual}"].font = tabela_titulo_font
    ws[f"A{linha_atual}"].fill = verde_escuro
    ws[f"A{linha_atual}"].alignment = centralizado
    
    # Informações da Conciliação
    # Total de Transações Conciliadas por Agrupamento
    ws.merge_cells(f"A{linha_atual + 1}:E{linha_atual + 1}")
    ws[f"A{linha_atual + 1}"].font = total_valor_font
    ws[f"A{linha_atual + 1}"].fill = verde_claro
    ws[f"A{linha_atual + 1}"].border = borda
    
    ws.merge_cells(f"F{linha_atual + 1}:J{linha_atual + 1}")
    ws[f"F{linha_atual + 1}"].font = total_valor_dados_font
    ws[f"F{linha_atual + 1}"].alignment = alinhamento_direita
    ws[f"F{linha_atual + 1}"].fill = verde_claro
    ws[f"F{linha_atual + 1}"].border = borda
    
    # Valor Total Conciliado por Agrupamento
    ws.merge_cells(f"A{linha_atual + 2}:E{linha_atual + 2}")
    ws[f"A{linha_atual + 2}"].font = total_valor_font
    ws[f"A{linha_atual + 2}"].fill = verde_claro
    ws[f"A{linha_atual + 2}"].border = borda
    
    ws.merge_cells(f"F{linha_atual + 2}:J{linha_atual + 2}")
    ws[f"F{linha_atual + 2}"].font = total_valor_dados_font
    ws[f"F{linha_atual + 2}"].alignment = alinhamento_direita
    ws[f"F{linha_atual + 2}"].number_format = formato_moeda
    ws[f"F{linha_atual + 2}"].fill = verde_claro
    ws[f"F{linha_atual + 2}"].border = borda
    
    # ==================================================
    # TABELA CONCILIADAS POR AGRUPAMENTO
    # ==================================================
    
    
    for cell in ws[f"A{linha_atual + 4}:J{linha_atual + 4}"][0]:
        cell.font = topo_tabela_font
        cell.fill = azul_escuro
        cell.alignment = centralizado
        cell.border = borda_branca
    
    ultima_linha = ws.max_row - 3
    
    for linha in range(linha_atual + 5, ultima_linha + 1):
        for cell in ws[f"A{linha}:J{linha}"][0]:
            cell.font = tabela_dados_font
            cell.border = borda
                
        ws[f"B{linha}"].alignment = alinhamento_direita
        ws[f"E{linha}"].alignment = alinhamento_direita
        ws[f"E{linha}"].number_format = formato_moeda
        ws[f"J{linha}"].alignment = alinhamento_direita
        ws[f"J{linha}"].number_format = formato_moeda       
    
    linha_total_agrupamentos = ws.max_row - 2
    
    for cell in ws[f"A{linha_total_agrupamentos}:J{linha_total_agrupamentos}"][0]:
        cell.fill = verde_claro
        cell.font = user_data_font
        cell.border = borda
    
    ws[f"D{linha_total_agrupamentos}"].alignment = alinhamento_direita
    ws[f"D{linha_total_agrupamentos}"].font = user_data_font
    ws[f"E{linha_total_agrupamentos}"].number_format = formato_moeda
    ws[f"E{linha_total_agrupamentos}"].alignment = alinhamento_direita
    ws[f"E{linha_total_agrupamentos}"].font = user_data_font
    ws[f"J{linha_total_agrupamentos}"].number_format = formato_moeda
    ws[f"J{linha_total_agrupamentos}"].alignment = alinhamento_direita
    ws[f"J{linha_total_agrupamentos}"].font = user_data_font
    
    linha_final = ws.max_row
    ws[f"A{linha_final}"].font = linha_final_font

def exportar_relatorio_excel_agrupamento(df_relatorio_conv, df_relatorio_div, df_relatorio_grafico):
    """
    Exporta o relatório para Excel.
    """

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Escreve os DataFrames
        
        df_relatorio_grafico.to_excel(
            writer,
            sheet_name="Resumo Executivo",
            index=False,
            header=False,
        )
        df_relatorio_conv.to_excel(
            writer,
            sheet_name="Transações Conciliadas",
            index=False,
            header=False,
        )

        df_relatorio_div.to_excel(
            writer,
            sheet_name="Transações Não Conciliadas",
            index=False,
            header=False,
        )

        # Obtém as abas
        ws_relatorio_grafico = writer.book["Resumo Executivo"]
        ws_conciliadas = writer.book["Transações Conciliadas"]
        ws_nao_conciliadas = writer.book["Transações Não Conciliadas"]

        # Aplica a formatação
        formatar_aba_grafico(ws_relatorio_grafico)
        formatar_aba_conciliadas_agrupamento(ws_conciliadas)
        formatar_aba_nao_conciliadas(ws_nao_conciliadas)
        
        # Ajuste personalizado das larguras
        largura_conciliadas = {"A": 45.86, "B": 21.29, "C": 43.43, "D": 43.43, "E": 16.29, "F": 25.57, "G": 27.14, "H": 57.00, "I": 57.00, "J": 21.00}
        largura_nao_conciliadas = {"A": 45.86, "B": 15.00, "C": 45.29, "D": 31.57, "E": 13.29, "F": 21.29}
        largura_relatorio_grafico = {"A": 31.86, "B": 16.29, "C": 18.57, "D": 18.57}
        altura_linhas = {1: 31.5, 2: 19.5, 7: 21.75, 8: 27.75}
        altura_grafico = {1: 60, 2: 21.75, 3: 7.5, 4: 15, 5: 15, 6:7.5, 7: 21.75, 8: 24, 9: 24, 10:24, 11: 24, 12: 24, 13: 24, 15: 20, 16: 20,
                                  17: 14.25, 18:14.25, 19:14.25, 20:14.25, 21:14.25, 22:14.25, 23:14.25, 24:14.25, 25:14.25, 26:14.25,
                                  27:14.25, 28:14.25, 29:14.25, 30:14.25, 31:15, 32:15, 33:14.25, 34:14.25}
        
        ajustar_largura_colunas(ws_conciliadas, largura_conciliadas)
        ajustar_largura_colunas(ws_nao_conciliadas, largura_nao_conciliadas)
        ajustar_altura_linhas(ws_conciliadas, altura_linhas)
        ajustar_altura_linhas(ws_nao_conciliadas, altura_linhas)
        ajustar_largura_colunas(ws_relatorio_grafico, largura_relatorio_grafico)
        ajustar_altura_linhas(ws_relatorio_grafico,altura_grafico)
        
    output.seek(0)
    return output.getvalue()


