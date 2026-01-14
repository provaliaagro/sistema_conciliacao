import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
import io
import funcoes_especificas as func
import streamlit as st

def criar_relatorio_conciliação(
    df_conciliado,
    saldo_inicial,  
    mov_extrato, 
    mov_controle,
    total_extrato,
    total_controle,
    nome_usuario
):
    """
    Cria um relatório completo da conciliação em um DataFrame estruturado
    que será exportado para excel
    """
    
    relatorio_dados = []
    
    # CABEÇALHO DO RELATÓRIO
    
    relatorio_dados.append(["RELATÓRIO FINAL DE CONCILIAÇÃO BANCÁRIA"])
    relatorio_dados.append([])  # Linha em branco
    relatorio_dados.append(["Usuário Responsável:", nome_usuario])
    relatorio_dados.append(["Data da Realização:", datetime.now().strftime("%d/%m/%Y %H:%M")])
    relatorio_dados.append([])  # Linha em branco
    
    # DADOS GERAIS DA CONCILIAÇÃO
    relatorio_dados.append(["DADOS GERAIS DA CONCILIAÇÃO"])
    relatorio_dados.append(["", "EXTRATO", "CONTROLE FINANCEIRO"])
    relatorio_dados.append(["Total de Movimentações:", mov_extrato, mov_controle])
    relatorio_dados.append(["Saldo Inicial (R$):", saldo_inicial, saldo_inicial])
    relatorio_dados.append(["Valor Total Movimentado (R$):", total_extrato, total_controle])
    relatorio_dados.append(["Saldo Final (R$):", st.session_state.saldo_final_ex, st.session_state.saldo_final_cf])
    relatorio_dados.append([])  # Linha em branco
    relatorio_dados.append([])  # Linha em branco
    
    # ESTATÍSTICAS DA CONCILIAÇÃO
    #conciliadas = len(df_conciliado[df_conciliado['status_conciliacao'] == 'CONCILIADA'])
    #nao_conciliadas = len(df_conciliado[df_conciliado['status_conciliacao'] == 'NÃO CONCILIADO'])
    
    #relatorio_dados.append(["ESTATÍSTICAS DA CONCILIAÇÃO"])
    #relatorio_dados.append(["Transações Conciliadas:", conciliadas])
    #relatorio_dados.append(["Transações Não Conciliadas:", nao_conciliadas])
    #relatorio_dados.append(["Taxa de Conciliação:", f"{(conciliadas/(conciliadas+nao_conciliadas)*100):.1f}%" if (conciliadas+nao_conciliadas) > 0 else "0%"])
    #relatorio_dados.append([])  # Linha em branco
    #relatorio_dados.append([])  # Linha em branco

    operacoes_divergentes = df_conciliado[df_conciliado['status_conciliacao'] == 'NÃO CONCILIADO']
    operacoes_convergentes = df_conciliado[df_conciliado['status_conciliacao'] == 'CONCILIADA']
    
    relatorio_div = relatorio_dados.copy()
    relatorio_conv = relatorio_dados.copy()
    
    # Relatório das Operações Não Conciliadas
    if len(operacoes_divergentes) > 0:
        relatorio_div.append(["OPERAÇÕES DIVERGENTES (NÃO CONCILIADAS)"])
        relatorio_div.append([])  # Linha em branco
        
        # Cabeçalho das operações divergentes
        cabecalho = [
            "Data", "Descrição", "Valor (R$)"
        ]
        
        # Adiciona dados que tem no extrato e não estão no controle financeiro
        relatorio_div.append(["Transações Não Conciliadas Presentes no Extrato:"])
        extrato_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_extrato,
                'descricao': row.get('descricao_extrato', ''),
                'valor': float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "vazio"
            }
            if linha['valor'] != "vazio":
                extrato_divergente.append(linha)
        
        extrato_divergente = func.ordenar_por_data_br(extrato_divergente)
        sum_divergente_extrato = len(extrato_divergente)
        total_divergente_extrato = 0
        for i in extrato_divergente:
            total_divergente_extrato += i['valor']
        
        relatorio_div.append(["Total de Transações Não Conciliadas:", sum_divergente_extrato])
        relatorio_div.append(["Valor Total das Transações Não Conciliadas (R$):", total_divergente_extrato])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho)
        
        for i in extrato_divergente:
            relatorio_div.append([i['data'], i['descricao'], i['valor']])
            
        relatorio_div.append([])
            
        # Adiciona dados que tem no controle financeiro e não estão no extrato
        relatorio_div.append(["Transações Não Conciliadas Presentes no Controle Financeiro:"])
        
        controle_divergente = []
        for _, row in operacoes_divergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = {
                'data': data_controle,
                'descricao': row.get('descricao_controle', ''),
                'valor': float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "vazio"
            }
            if linha['valor'] != "vazio":
                controle_divergente.append(linha)
            
        controle_divergente = func.ordenar_por_data_br(controle_divergente)
        sum_divergente_controle = len(controle_divergente)
        total_divergente_controle = 0
        for i in controle_divergente:
            total_divergente_controle += i['valor']
    
        relatorio_div.append(["Total de Transações Não Conciliadas:", sum_divergente_controle])
        relatorio_div.append(["Valor Total das Transações Não Conciliadas (R$):", total_divergente_controle])
        relatorio_div.append([])  # Linha em branco
        relatorio_div.append(cabecalho)
        
        for i in controle_divergente:
                relatorio_div.append([i['data'], i['descricao'], i['valor']])
            
    else:
        relatorio_div.append(["NENHUMA OPERAÇÃO DIVERGENTE ENCONTRADA"])
        
    # Relatório das Operações Conciliadas
    if len(operacoes_convergentes) > 0:
        sum_convergentes = len(operacoes_convergentes)
        total_convergentes = sum(operacoes_convergentes['valor_extrato'])
        
        relatorio_conv.append(["OPERAÇÕES CONVERGENTES (CONCILIADAS)"])
        relatorio_conv.append(["Total de Transações Conciliadas:", sum_convergentes])
        relatorio_conv.append(["Valor Total Conciliado (R$):", total_convergentes])
        relatorio_conv.append([])  # Linha em branco
        
        # Cabeçalho das operações convergentes
        cabecalho = [
            "Data Extrato", "Descrição Extrato", "Valor Extrato (R$)", 
            "Data Controle", "Descrição Controle", "Valor Controle (R$)"
        ]
        relatorio_conv.append(cabecalho)
        
        # Adiciona cada operação convergente
        for _, row in operacoes_convergentes.iterrows():
            # Verifica se tem dados de data (pode não ter se for merge outer)
            data_extrato = row.get('data_extrato', '') if 'data_extrato' in row else ''
            data_controle = row.get('data_controle', '') if 'data_controle' in row else ''
            
            linha = [
                data_extrato,
                row.get('descricao_extrato', ''),
                float(row.get('valor_extrato', 0)) if pd.notna(row.get('valor_extrato')) else "",
                data_controle,
                row.get('descricao_controle', ''),
                float(row.get('valor_controle', 0)) if pd.notna(row.get('valor_controle')) else "",
            ]
            relatorio_conv.append(linha)
    else:
        relatorio_conv.append(["NENHUMA OPERAÇÃO CONCILIADA ENCONTRADA"])
    
    # Converte para DataFrame
    df_relatorio_div = pd.DataFrame(relatorio_div)
    df_relatorio_conv = pd.DataFrame(relatorio_conv)
    
    return df_relatorio_conv, df_relatorio_div


def exportar_relatorio_excel(df_relatorio_conv, df_relatorio_div):
    """
    Exporta o DataFrame do relatório para Excel em memória
    """    
    
    # Cria um buffer em memória
    output = io.BytesIO()
    
    # Cria o Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Aba com os Convergentes
        df_relatorio_conv.to_excel(
            writer,
            index=False,
            header=False,  
            sheet_name='Transações Conciliadas'
        )
                
        # Aba com os Divergentes
        df_relatorio_div.to_excel(
            writer,
            index=False,
            header=False,  
            sheet_name='Transações Não Conciliadas'
        )
        
        # CRIAÇÃO DO ESTILO PERSONALIZADO
        workbook = writer.book
        
        # 1. Fontes
        fonte_titulo = Font(name='Calibri', size=14, bold=True, color='000000')
        fonte_cabecalho = Font(name='Calibri', size=11, bold=True, color='000000')
        fonte_normal = Font(name='Calibri', size=10)
        fonte_negativo = Font(name='Calibri', size=10, color='FF0000')
        fonte_positivo = Font(name='Calibri', size=10, color='0000FF')
        
        # 2. Cores
        cor_titulo_conciliado = PatternFill(start_color='5fb7fa', end_color='5fb7fa', fill_type='solid')
        cor_cabecalho_conciliado = PatternFill(start_color='5fb7fa', end_color='5fb7fa', fill_type='solid')
        cor_linha_par = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cor_titulo_divergente = PatternFill(start_color='fc7474', end_color='fc7474', fill_type='solid')
        cor_cabecalho_divergente = PatternFill(start_color='fc7474', end_color='fc7474', fill_type='solid')
        
        # 3. Borda
        borda_fina = Side(border_style='thin', color='D0D0D0')
        borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

        # 4. Alinhamento
        alinhamento_centro = Alignment(horizontal='center', vertical='center')
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
        alinhamento_direita = Alignment(horizontal='right', vertical='center')
        
        # 5. Formato do Número
        formato_numero = "#,##0.00"
        
        # Aplica estilos em ambas as abas
        for sheet_name in ['Transações Conciliadas', 'Transações Não Conciliadas']:
            worksheet = workbook[sheet_name]

            # PRIMEIRO: Aplica cor de fundo para linhas pares
            for i, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
                if i > 17 and i % 2 == 0:  # Começa após os cabeçalhos
                    for cell in row:
                        cell.fill = cor_linha_par
            
            # SEGUNDO: Itera por todas as células aplicando formatação
            for row in worksheet.iter_rows():
                for cell in row:
                    # Aplica borda para todas as células
                    cell.border = borda
                    
                    # Para células NÃO numéricas
                    if not isinstance(cell.value, (int, float)):
                        # Aplica fonte normal e alinhamento
                        cell.font = fonte_normal
                        cell.alignment = alinhamento_esquerda                            
                    else :
                        # Para células numéricas
                        cell.alignment = alinhamento_direita
                        
                        # Formatando o cabeçalho
                        cabecalho_numerico = [9, 10, 11, 16, 18]
                        if cell.row in cabecalho_numerico and cell.column_letter in ["B", "C"]:
                            cell.number_format = formato_numero
                            if cell.value < 0:
                                cell.font = fonte_negativo  # Vermelho para negativos
                            elif cell.value > 0:
                                cell.font = fonte_positivo  # Azul para positivos
                            else:
                                cell.font = fonte_normal  # Zero mantém fonte normal
                        
                        
                        # Formatando as demais colunas
                        if cell.row >= 19 and cell.column_letter in ["B", "C", "F"]:
                            cell.number_format = formato_numero
                            
                            if cell.value < 0:
                                cell.font = fonte_negativo  # Vermelho para negativos
                            elif cell.value > 0:
                                cell.font = fonte_positivo  # Azul para positivos
                            else:
                                cell.font = fonte_normal  # Zero mantém fonte normal
                    
                    # Estilo para títulos (linhas com texto em negrito)
                    if sheet_name == "Transações Conciliadas":
                        if cell.value and isinstance(cell.value, str):
                            
                            if any(titulo in cell.value for titulo in [
                                "RELATÓRIO FINAL DE CONCILIAÇÃO BANCÁRIA",
                                "DADOS GERAIS DA CONCILIAÇÃO", 
                                "RESUMO DE MOVIMENTAÇÕES",
                                "OPERAÇÕES CONVERGENTES",
                            ]):
                                cell.font = fonte_titulo
                                cell.fill = cor_titulo_conciliado
                                cell.alignment = alinhamento_centro
                            
                            # Estilo para cabeçalhos de tabela
                            elif cell.value in ["EXTRATO","CONTROLE FINANCEIRO","Data Extrato", "Descrição Extrato", "Valor Extrato (R$)", 
                                            "Data Controle", "Descrição Controle", "Valor Controle (R$)"]:
                                cell.font = fonte_cabecalho
                                cell.fill = cor_cabecalho_conciliado
                                cell.alignment = alinhamento_centro
                                
                    elif sheet_name == "Transações Não Conciliadas":
                        if cell.value and isinstance(cell.value, str):
                            
                            if any(titulo in cell.value for titulo in [
                                "RELATÓRIO FINAL DE CONCILIAÇÃO BANCÁRIA",
                                "DADOS GERAIS DA CONCILIAÇÃO", 
                                "RESUMO DE MOVIMENTAÇÕES",
                                "OPERAÇÕES DIVERGENTES",
                                "Transações Não Conciliadas Presentes no Extrato:",
                                "Transações Não Conciliadas Presentes no Controle Financeiro:"
                            ]):
                                cell.font = fonte_titulo
                                cell.fill = cor_titulo_divergente
                                cell.alignment = alinhamento_centro
                            
                            # Estilo para cabeçalhos de tabela
                            elif cell.value in ["EXTRATO", "CONTROLE FINANCEIRO", "Data", "Descrição", "Valor (R$)"]:
                                cell.font = fonte_cabecalho
                                cell.fill = cor_cabecalho_divergente
                                cell.alignment = alinhamento_centro
            
        
        # Ajusta a largura das colunas
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Pega os bytes
    excel_bytes = output.getvalue()
    output.close()
    
    return excel_bytes
